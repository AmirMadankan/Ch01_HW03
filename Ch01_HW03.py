import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

ws = load_workbook("files/Ch01_HW03_linear_regression_data.xlsx")["Sheet1"]

n = ws.max_row + 1
x = []
y = []
for i in range(2, n):
    x.append([ws[f"A{i}"].value, ws[f"B{i}"].value, ws[f"C{i}"].value])
    y.append([ws[f"D{i}"].value])

x = np.array(x).reshape(-1, 3)
y = np.array(y).reshape(-1, 1)
